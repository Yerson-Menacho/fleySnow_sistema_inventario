from flask import Blueprint, render_template, request, jsonify, send_file
from io import BytesIO
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import get_db_connection
from app.routes.reportes.reportes_usuarios import reportes_bp
from app.models.movimientos_model import listar_notas_movimiento
from app.utils.auth import permiso_de_admin


# ==============================
# 📋 Vista principal - Reporte de Notas de Movimiento
# ==============================
@reportes_bp.route('/reportes/movimientos')
@permiso_de_admin('movimientos.ver')
def vista_reporte_movimientos():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # 🔹 Cargar los tipos de movimiento
    cursor.execute("SELECT id_tipo, nombre FROM tipo_movimiento ORDER BY nombre ASC")
    tipos_movimiento = cursor.fetchall()

    cursor.close()
    conn.close()

    # 🔹 Pasar la lista al template
    return render_template('reportes/reporte_movimiento.html', tipos_movimiento=tipos_movimiento)

# ==============================
# 📊 Data vía AJAX - Reporte Notas de Movimiento
# ==============================
@reportes_bp.route('/reportes/movimientos/data')
@permiso_de_admin('movimientos.ver')
def data_reporte_movimientos():
    p_q = request.args.get("q", None)
    p_id_tipo = request.args.get("id_tipo", None)
    p_estado = request.args.get("estado", None)
    p_fecha_inicio = request.args.get("fecha_inicio", None)
    p_fecha_fin = request.args.get("fecha_fin", None)

    # Convertir id_tipo a entero solo si tiene valor
    p_id_tipo = int(p_id_tipo) if p_id_tipo else None

    data_raw = listar_notas_movimiento(p_q, p_id_tipo, p_estado, p_fecha_inicio, p_fecha_fin)

    data = []
    for n in data_raw:
        data.append({
            "id_nota": n["id_nota"],
            "codigo_nota": n["codigo_nota"],
            "tipo_movimiento": n.get("tipo_movimiento", ""),
            "fecha": n.get("fecha", ""),
            "referencia": n.get("referencia", ""),
            "usuario": n.get("usuario", ""),
            "origen_destino": n.get("origen_destino", ""),
            "observacion": n.get("observacion", ""),
            "estado": n.get("estado", ""),
        })

    return jsonify(data)


# ==============================
# 📈 Exportar Excel - Reporte Notas de Movimiento
# ==============================
@reportes_bp.route('/reportes/movimientos/excel')
@permiso_de_admin('movimientos.ver')
def exportar_reporte_movimientos_excel():
    movimientos = listar_notas_movimiento()

    # ✅ Reducir tamaño del ID (últimos 4 dígitos)
    for m in movimientos:
        if m.get("id_nota"):
            m["id_nota"] = str(m["id_nota"])[-4:]
        else:
            m["id_nota"] = ""

    df = pd.DataFrame(movimientos)

    if not df.empty:
        df = df.rename(columns={
            "id_nota": "ID",
            "codigo_nota": "Código",
            "tipo_movimiento": "Tipo",
            "fecha": "Fecha",
            "referencia": "Referencia",
            "usuario": "Responsable",
            "origen_destino": "Origen/Destino",
            "observacion": "Observaciones",
            "estado": "Estado"
        })

        df = df[[
            "ID", "Código", "Tipo", "Fecha", "Referencia",
            "Responsable", "Origen/Destino", "Observaciones", "Estado"
        ]]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='NotasMovimiento')
        wb = writer.book
        ws = wb.active

        # === CABECERA EMPRESARIAL ===
        ws.insert_rows(1, 6)
        ws.merge_cells('A1:I1')
        ws.merge_cells('A2:I2')
        ws.merge_cells('A3:I3')
        ws.merge_cells('A4:I4')
        ws.merge_cells('A5:I5')

        ws['A1'] = "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L."
        ws['A2'] = "RUC: 20610415726"
        ws['A3'] = "Dirección: Otr. Sector Conchuc Lote 7, Caraz, Huaylas, Ancash, Perú"
        ws['A4'] = "📑 Reporte de Notas de Movimiento"
        ws['A5'] = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        for i in range(1, 6):
            cell = ws[f"A{i}"]
            cell.font = Font(bold=True, size=12 if i == 1 else 10, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # === ENCABEZADO ===
        header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[7]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # === CUERPO ===
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # === AUTO AJUSTE ===
        for i, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            column = get_column_letter(i)
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_movimientos.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==============================
# 📄 Exportar PDF - Reporte Notas de Movimiento
# ==============================
@reportes_bp.route('/reportes/movimientos/pdf')
@permiso_de_admin('movimientos.ver')
def exportar_reporte_movimientos_pdf():
    movimientos = listar_notas_movimiento()
    output = BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    width, height = letter

    # === CABECERA EMPRESARIAL ===
    logo_path = "app/static/img/logo.png"

    # Fondo superior verde institucional
    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(0, height - 1.3 * inch, width, 1.3 * inch, fill=True, stroke=False)

    # Logo (si existe)
    try:
        p.drawImage(logo_path, 0.6 * inch, height - 1.1 * inch, width=70, height=70, mask='auto')
    except:
        pass

    # Texto institucional
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 13)
    p.drawString(1.8 * inch, height - 0.7 * inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L.")
    p.setFont("Helvetica", 9)
    p.drawString(1.8 * inch, height - 0.9 * inch, "RUC: 20610415726")
    p.drawString(1.8 * inch, height - 1.05 * inch, "Dirección: Otr. Sector Conchuc Lote 7, Caraz, Huaylas, Ancash, Perú")

    # Línea separadora
    p.setStrokeColorRGB(0.18, 0.49, 0.20)
    p.setLineWidth(2)
    p.line(0.6 * inch, height - 1.4 * inch, width - 0.6 * inch, height - 1.4 * inch)

    # === TÍTULO DEL REPORTE ===
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(1 * inch, height - 1.8 * inch, "📑 Reporte de Notas de Movimiento")

    # Fecha de generación
    p.setFont("Helvetica", 9)
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    p.drawRightString(width - 0.6 * inch, height - 1.8 * inch, f"Generado el: {fecha_actual}")

    # === ENCABEZADO DE TABLA ===
    y = height - 2.2 * inch
    header_x = 0.4 * inch
    header_w = width - 0.8 * inch
    header_h = 0.3 * inch

    columnas = ["ID", "Código", "Tipo", "Fecha", "Referencia", "Responsable", "Origen/Destino", "Observaciones", "Estado"]
    proporciones = [0.04, 0.09, 0.07, 0.07, 0.09, 0.15, 0.18, 0.20, 0.08]  # Suma ≈ 1.0
    anchos = [header_w * p for p in proporciones]

    # Fondo del encabezado
    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(header_x, y - 0.2 * inch, header_w, header_h, fill=True, stroke=False)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 8)

    # Nombres de columnas centradas
    x_actual = header_x
    for i, col in enumerate(columnas):
        p.drawCentredString(x_actual + anchos[i] / 2, y - 0.05 * inch, col)
        x_actual += anchos[i]

    # === CUERPO DE TABLA ===
    p.setFont("Helvetica", 7.5)
    p.setFillColor(colors.black)
    y -= 0.35 * inch

    for m in movimientos:
        if y < 1 * inch:
            # Nueva página
            p.showPage()
            y = height - 1 * inch

            # Repetir encabezado
            p.setFont("Helvetica-Bold", 8)
            p.setFillColorRGB(0.18, 0.49, 0.20)
            p.rect(header_x, y - 0.2 * inch, header_w, header_h, fill=True, stroke=False)
            p.setFillColor(colors.white)
            x_actual = header_x
            for i, col in enumerate(columnas):
                p.drawCentredString(x_actual + anchos[i] / 2, y - 0.05 * inch, col)
                x_actual += anchos[i]
            y -= 0.35 * inch
            p.setFont("Helvetica", 7.5)
            p.setFillColor(colors.black)

        datos = [
            str(m.get("id_nota", "")),
            str(m.get("codigo_nota", "")),
            str(m.get("tipo_movimiento", "")),
            str(m.get("fecha", ""))[:10],
            str(m.get("referencia", ""))[:20],
            str(m.get("usuario", ""))[:20],
            str(m.get("origen_destino", ""))[:20],
            str(m.get("observacion", ""))[:25],
            str(m.get("estado", "")),
        ]

        x_actual = header_x
        for i, dato in enumerate(datos):
            p.drawCentredString(x_actual + anchos[i] / 2, y, dato)
            x_actual += anchos[i]
        y -= 0.25 * inch

    # === PIE DE PÁGINA ===
    p.setFont("Helvetica-Oblique", 8)
    p.setFillColor(colors.gray)
    p.drawCentredString(width / 2, 0.5 * inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L. - Sistema de Gestión © 2025")

    p.save()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_movimientos.pdf',
        mimetype='application/pdf'
    )
