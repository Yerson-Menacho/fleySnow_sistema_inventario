import pandas as pd
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.routes import reportes
from flask import Blueprint, jsonify, render_template, flash, request, send_file
from app.models.insumos_model import listar_insumos_stock_bajo
from app.models.reportes_model import listar_agricultores_reporte
from app.utils.auth import permiso_de_admin 

from app.utils.auth import permiso_de_admin
from app.routes.reportes.reportes_usuarios import reportes_bp


# ==============================
# 👨‍🌾 Reporte: Agricultores
# ==============================
@reportes_bp.route('/reportes/agricultores')
@permiso_de_admin('agricultores.ver')
def vista_reporte_agricultores():
    return render_template('reportes/reporte_agricultores.html')


# ==============================
# 📊 Data vía AJAX - Reporte agricultores
# ==============================
@reportes_bp.route('/reportes/agricultores/data')
@permiso_de_admin('agricultores.ver')
def data_reporte_agricultores():
    def norm(v):
        if v is None:
            return None
        v = v.strip()
        return v if v not in ("", "todos", "TODOS") else None

    filtros = {
        "q": norm(request.args.get("q")),
        "estado": norm(request.args.get("estado")),
        "fecha_inicio": norm(request.args.get("fecha_inicio")),
        "fecha_fin": norm(request.args.get("fecha_fin")),
    }

    data = listar_agricultores_reporte(filtros)

    # ✅ Asegurarse de que siempre sea una lista
    if isinstance(data, dict):
        data = [data]

    return jsonify(data)

# ==============================
# 📈 Exportar Excel Agricultores
# ==============================
@reportes_bp.route('/reportes/agricultores/excel')
@permiso_de_admin('agricultores.ver')
def exportar_reporte_agricultores_excel():
    filtros = request.args.to_dict()
    agricultores = listar_agricultores_reporte(filtros)

    df = pd.DataFrame(agricultores)

    if not df.empty:
        df.rename(columns={
            "id_agricultor": "ID",
            "codigo_agricultor": "Código",
            "nombre_completo": "Nombre completo",
            "dni": "DNI",
            "telefono": "Teléfono",
            "correo": "Correo",
            "direccion": "Dirección",
            "zona": "Zona",
            "cultivo_principal": "Cultivo principal",
            "fecha_registro": "Fecha registro",
            "estado": "Estado"
        }, inplace=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Agricultores')

        wb = writer.book
        ws = wb.active

        # === Encabezado Empresarial ===
        ws.insert_rows(1, 6)
        ws.merge_cells('A1:L1')
        ws.merge_cells('A2:L2')
        ws.merge_cells('A3:L3')
        ws.merge_cells('A4:L4')
        ws.merge_cells('A5:L5')

        ws['A1'] = "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L."
        ws['A2'] = "RUC: 20610415726"
        ws['A3'] = "Dirección Legal: Otr. Sector Conchuc Lote 7, Caraz, Huaylas, Ancash, Perú"
        ws['A4'] = ""
        ws['A5'] = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        for i in range(1, 6):
            cell = ws[f"A{i}"]
            cell.font = Font(bold=True, size=11 if i == 1 else 10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # === Encabezado Tabla ===
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )

        for cell in ws[7]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # === Bordes del cuerpo ===
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Ajuste automático de ancho
        for i, column in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].auto_size = True

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_agricultores.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ==============================
# 📄 Exportar PDF Agricultores
# ==============================
@reportes_bp.route('/reportes/agricultores/pdf')
@permiso_de_admin('agricultores.ver')
def exportar_reporte_agricultores_pdf():
    filtros = request.args.to_dict()
    agricultores = listar_agricultores_reporte(filtros)

    output = BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    width, height = letter

    # === Cabecera Empresarial ===
    logo_path = "app/static/img/logo.png"
    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(0, height - 1.3*inch, width, 1.3*inch, fill=True, stroke=False)

    try:
        p.drawImage(logo_path, 0.6*inch, height - 1.1*inch, width=70, height=70, mask='auto')
    except: pass

    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 13)
    p.drawString(1.8*inch, height - 0.7*inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L.")
    p.setFont("Helvetica", 9)
    p.drawString(1.8*inch, height - 0.9*inch, "RUC: 20610415726")
    p.drawString(1.8*inch, height - 1.05*inch, "Dirección: Otr. Sector Conchuc Lote 7, Caraz, Huaylas, Ancash, Perú")

    p.setStrokeColorRGB(0.18, 0.49, 0.20)
    p.setLineWidth(2)
    p.line(0.6*inch, height - 1.4*inch, width - 0.6*inch, height - 1.4*inch)

    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(1*inch, height - 1.8*inch, "📋 Reporte de Agricultores")

    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    p.setFont("Helvetica", 9)
    p.drawRightString(width - 0.6*inch, height - 1.8*inch, f"Generado el: {fecha_actual}")

    # === Encabezado de tabla ===
    y = height - 2.2*inch
    header_x = 0.4*inch
    header_w = width - 0.8*inch
    header_h = 0.3*inch

    columnas = ["Código", "Nombre", "DNI", "Teléfono", "Correo", "Dirección", "Zona", "Cultivo", "Estado", "Registro"]
    proporciones = [0.08,0.18,0.08,0.09,0.15,0.15,0.08,0.08,0.06,0.05]
    anchos = [header_w * p for p in proporciones]

    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(header_x, y - 0.2*inch, header_w, header_h, fill=True, stroke=False)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 8)

    x_actual = header_x
    for i, col in enumerate(columnas):
        p.drawCentredString(x_actual + anchos[i]/2, y - 0.05*inch, col)
        x_actual += anchos[i]

    # === Cuerpo ===
    p.setFont("Helvetica", 7.5)
    p.setFillColor(colors.black)
    y -= 0.35*inch

    for a in agricultores:
        if y < 1*inch:
            p.showPage()
            y = height - 1*inch
            p.setFont("Helvetica-Bold", 8)
            p.setFillColorRGB(0.18, 0.49, 0.20)
            p.rect(header_x, y - 0.2*inch, header_w, header_h, fill=True, stroke=False)
            p.setFillColor(colors.white)
            x_actual = header_x
            for i, col in enumerate(columnas):
                p.drawCentredString(x_actual + anchos[i]/2, y - 0.05*inch, col)
                x_actual += anchos[i]
            y -= 0.35*inch
            p.setFont("Helvetica", 7.5)
            p.setFillColor(colors.black)

        datos = [
            str(a.get("codigo_agricultor", "")),
            (a.get("nombre_completo") or "")[:20],
            str(a.get("dni", "")),
            str(a.get("telefono", "")),
            (a.get("correo") or "")[:25],
            (a.get("direccion") or "")[:20],
            (a.get("zona") or "")[:10],
            (a.get("cultivo_principal") or "")[:10],
            str(a.get("estado") or ""),
            str(a.get("fecha_registro") or "")
        ]

        x_actual = header_x
        for i, dato in enumerate(datos):
            p.drawCentredString(x_actual + anchos[i]/2, y, dato)
            x_actual += anchos[i]

        y -= 0.25*inch

    # Pie de página
    p.setFont("Helvetica-Oblique", 8)
    p.setFillColor(colors.gray)
    p.drawCentredString(width/2, 0.5*inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L. - Sistema de Gestión © 2025")

    p.save()
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_agricultores.pdf',
        mimetype='application/pdf'
    )

