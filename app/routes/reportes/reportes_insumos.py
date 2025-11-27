from flask import Blueprint, render_template, request, jsonify, send_file
from io import BytesIO
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.routes.reportes.reportes_usuarios import reportes_bp
from app.models.categorias_model import listar_categorias
from app.models.reportes_model import listar_insumos_reporte
from app.utils.auth import permiso_de_admin

# ==============================
# 📋 Vista principal - Reporte Insumos
# ==============================
@reportes_bp.route('/reportes/insumos')
@permiso_de_admin('insumos.ver')
def vista_reporte_insumos():
    categorias = listar_categorias()
    return render_template('reportes/reporte_insumos.html', categorias=categorias)

# ==============================
# 📊 Data vía AJAX - Reporte Insumos
# ==============================
@reportes_bp.route('/reportes/insumos/data')
@permiso_de_admin('insumos.ver')
def data_reporte_insumos():
    def norm(v):
        if v is None:
            return None
        v = v.strip()
        return v if v not in ("", "todos", "TODOS") else None

    filtros = {
        "q": norm(request.args.get("q")),
        "id_categoria": norm(request.args.get("id_categoria")),   "id_variedad": norm(request.args.get("id_variedad")),
        "estado": norm(request.args.get("estado")), "fecha_inicio": norm(request.args.get("fecha_inicio")),
        "fecha_fin": norm(request.args.get("fecha_fin")),
    }

    data_raw = listar_insumos_reporte(filtros)
    if isinstance(data_raw, dict):
        data_raw = [data_raw]

    data = []
    for i in data_raw:
        data.append({
            "id": i["id_insumo"], "codigo": i["codigo_insumo"],  "nombre": i["nombre_insumo"], "categoria": i.get("categoria", ""),
            "variedad": i.get("variedad", ""),  "stock": i.get("stock_actual", 0),  "unidad": i.get("unidad_medida", ""),
            "estado": "Activo" if i.get("estado") == 1 else "Inactivo", "fecha_creacion": i.get("fecha_ingreso"),
        })

    return jsonify(data)


# ==============================
# 📈 Exportar Excel
# ==============================
@reportes_bp.route('/reportes/insumos/excel')
@permiso_de_admin('insumos.ver')
def exportar_reporte_insumos_excel():
    filtros = request.args.to_dict()
    insumos = listar_insumos_reporte(filtros)

    df = pd.DataFrame(insumos)
    if not df.empty:
        df.rename(columns={
            "id_insumo": "ID",
            "codigo_insumo": "Código",
            "nombre_insumo": "Nombre",
            "categoria": "Categoría",
            "variedad": "Variedad",
            "stock_actual": "Stock Actual",
            "stock_minimo": "Stock Mínimo",
            "unidad_medida": "Unidad",
            "fecha_ingreso": "Ingreso",
            "fecha_vencimiento": "Vencimiento",
            "descripcion": "Descripción",
            "estado": "Estado"
        }, inplace=True)

        # Estado 1 = Activo / 0 = Descontinuado
        df["Estado"] = df["Estado"].apply(lambda x: "Activo" if x == 1 else "Descontinuado")

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Insumos')
        wb = writer.book
        ws = wb.active

        # === ENCABEZADO EMPRESARIAL ===
        ws.insert_rows(1, 6)
        ws.merge_cells('A1:L1')
        ws.merge_cells('A2:L2')
        ws.merge_cells('A3:L3')
        ws.merge_cells('A4:L4')
        ws.merge_cells('A5:L5')

        ws['A1'] = "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L."
        ws['A2'] = "RUC: 20610415726"
        ws['A3'] = "Dirección Legal: Otr. Sector Conchuc Lote 7, Caraz, Huaylas, Ancash, Perú"
        ws['A4'] = "Reporte de Insumos del Sistema"
        ws['A5'] = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        for i in range(1, 6):
            cell = ws[f"A{i}"]
            cell.font = Font(bold=True, size=11 if i == 1 else 10, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # === ENCABEZADO DE TABLA ===
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )

        for cell in ws[7]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # === CUERPO DE TABLA ===
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # === AJUSTE AUTOMÁTICO DE ANCHO ===
        for i, column in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].auto_size = True

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_insumos.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ==============================
# 📄 Exportar PDF
# ==============================
@reportes_bp.route('/reportes/insumos/pdf')
@permiso_de_admin('insumos.ver')
def exportar_reporte_insumos_pdf():
    filtros = request.args.to_dict()
    insumos = listar_insumos_reporte(filtros)

    output = BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    width, height = letter

    # === CABECERA EMPRESARIAL ===
    logo_path = "app/static/img/logo.png"
    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(0, height - 1.3 * inch, width, 1.3 * inch, fill=True, stroke=False)

    try:
        p.drawImage(logo_path, 0.6 * inch, height - 1.1 * inch, width=70, height=70, mask='auto')
    except:
        pass

    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 13)
    p.drawString(1.8 * inch, height - 0.7 * inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L.")

    p.setFont("Helvetica", 9)
    p.drawString(1.8 * inch, height - 0.9 * inch, "RUC: 20610415726")
    p.drawString(1.8 * inch, height - 1.05 * inch, "Dirección: Caraz, Huaylas, Ancash, Perú")

    # === TÍTULO DEL REPORTE ===
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(1 * inch, height - 1.8 * inch, "📦 Reporte de Insumos del Sistema")

    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    p.setFont("Helvetica", 9)
    p.drawRightString(width - 0.6 * inch, height - 1.8 * inch, f"Generado el: {fecha_actual}")

    # === ENCABEZADO DE TABLA ===
    y = height - 2.2 * inch
    columnas = ["ID", "Código", "Nombre", "Categoría", "Stock", "Unidad", "Estado"]
    posiciones = [0.6, 1.0, 1.8, 3.3, 5.0, 5.8, 6.6]

    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(0.5 * inch, y - 0.2 * inch, width - 1.0 * inch, 0.3 * inch, fill=True, stroke=False)
    p.setFont("Helvetica-Bold", 9)
    p.setFillColor(colors.white)
    for i, col in enumerate(columnas):
        p.drawString(posiciones[i] * inch, y - 0.05 * inch, col)

    # === CUERPO DE TABLA ===
    p.setFont("Helvetica", 8)
    p.setFillColor(colors.black)
    y -= 0.35 * inch

    for ins in insumos:
        if y < 1 * inch:
            p.showPage()
            y = height - 1 * inch
            p.setFont("Helvetica-Bold", 9)
            p.setFillColorRGB(0.18, 0.49, 0.20)
            p.rect(0.5 * inch, y - 0.2 * inch, width - 1.0 * inch, 0.3 * inch, fill=True, stroke=False)
            p.setFillColor(colors.white)
            for i, col in enumerate(columnas):
                p.drawString(posiciones[i] * inch, y - 0.05 * inch, col)
            y -= 0.35 * inch
            p.setFont("Helvetica", 8)
            p.setFillColor(colors.black)

        estado_texto = "Activo" if ins.get("estado") == 1 else "Descontinuado"

        datos = [
            str(ins.get("id_insumo", "")),
            str(ins.get("codigo_insumo", "")),
            (ins.get("nombre_insumo") or "")[:20],
            str(ins.get("categoria", "")),
            str(ins.get("stock_actual", "")),
            str(ins.get("unidad_medida", "")),
            estado_texto,
        ]

        for i, dato in enumerate(datos):
            p.drawString(posiciones[i] * inch, y, dato)

        y -= 0.25 * inch

    # Pie de página
    p.setFont("Helvetica-Oblique", 8)
    p.setFillColor(colors.gray)
    p.drawCentredString(width / 2, 0.5 * inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L. - Sistema de Gestión © 2025")

    p.save()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_insumos.pdf',
        mimetype='application/pdf'
    )
