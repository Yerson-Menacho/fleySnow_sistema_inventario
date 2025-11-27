import pandas as pd
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask import Blueprint, jsonify, render_template, flash, request, send_file
from app.models.insumos_model import listar_insumos_stock_bajo
from app.models.reportes_model import listar_roles, listar_usuarios_reporte
from app.utils.auth import permiso_de_admin 

reportes_bp = Blueprint('reportes', __name__)

# ==============================
# 📌 Reporte: Insumos con stock bajo
# ==============================
@reportes_bp.route('/reportes/stock-bajo')
@permiso_de_admin('stock.ver')
def stock_bajo():
    insumos = listar_insumos_stock_bajo()

    if not insumos:
        flash("📉 No hay insumos con stock bajo en este momento.", "info")

    return render_template(
        'reportes/stock_bajo.html',
        insumos=insumos
    )

# ==============================
# 👤 Reporte: Usuarios
# ==============================
@reportes_bp.route('/reportes/usuarios')
@permiso_de_admin('usuarios.ver')
def vista_reporte_usuarios():
    roles = listar_roles()
    return render_template('reportes/reporte_usuarios.html', roles=roles)


# ==============================
# 📊 Data vía AJAX - Reporte usuario
# ==============================
@reportes_bp.route('/reportes/usuarios/data')
@permiso_de_admin('usuarios.ver')
def data_reporte_usuarios():
    def norm(v):
        if v is None:
            return None
        v = v.strip()
        return v if v not in ("", "todos", "TODOS") else None

    filtros = {
        "q": norm(request.args.get("q")),
        "id_rol": norm(request.args.get("id_rol")),
        "estado": norm(request.args.get("estado")),
        "fecha_inicio": norm(request.args.get("fecha_inicio")),
        "fecha_fin": norm(request.args.get("fecha_fin")),
    }

    data = listar_usuarios_reporte(filtros)

    # Asegurarse de que siempre sea una lista
    if isinstance(data, dict):
        data = [data]

    return jsonify(data)

# ==============================
# 📈 Exportar Excel
# ==============================
@reportes_bp.route('/reportes/usuarios/excel')
@permiso_de_admin('usuarios.ver')
def exportar_reporte_usuarios_excel():
    filtros = request.args.to_dict()
    usuarios = listar_usuarios_reporte(filtros)

    # Crear DataFrame
    df = pd.DataFrame(usuarios)

    if not df.empty:
        df.rename(columns={
            "id_usuario": "ID",
            "nombre_completo": "Nombre",
            "dni": "DNI",
            "nombre_rol": "Rol",
            "email": "Email",
            "estado": "Estado",
            "fecha_creacion": "Creación"
        }, inplace=True)

        # Reemplazar valores numéricos por texto en Estado
        df["Estado"] = df["Estado"].apply(lambda x: "Activo" if x == 1 else "Inactivo")

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Usuarios')

        wb = writer.book
        ws = wb.active

        # === ENCABEZADO EMPRESARIAL ===
        ws.insert_rows(1, 6)
        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')
        ws.merge_cells('A3:G3')
        ws.merge_cells('A4:G4')
        ws.merge_cells('A5:G5')

        ws['A1'] = "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L."
        ws['A2'] = "RUC: 20610415726"
        ws['A3'] = "Dirección Legal: Otr. Sector Conchuc Lote. 7 Sec. Conchuc (Altura Cruce Trapiche Carretera)"
        ws['A4'] = "Distrito / Ciudad: Caraz | Provincia: Huaylas | Departamento: Ancash, Perú"
        ws['A5'] = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        # Estilo encabezado de empresa
        for i in range(1, 6):
            cell = ws[f"A{i}"]
            cell.font = Font(bold=True, size=11 if i == 1 else 10, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 25

        # === ENCABEZADO DE TABLA ===
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # verde oscuro
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )

        for cell in ws[7]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # === CUERPO DE LA TABLA ===
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # === AJUSTE AUTOMÁTICO DE ANCHO ===
        for i, column in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].auto_size = True

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_usuarios.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ==============================
# 📄 Exportar PDF
# ==============================
@reportes_bp.route('/reportes/usuarios/pdf')
@permiso_de_admin('usuarios.ver')
def exportar_reporte_usuarios_pdf():
    filtros = request.args.to_dict()
    usuarios = listar_usuarios_reporte(filtros)

    output = BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    width, height = letter

    # === CABECERA EMPRESARIAL ===
    logo_path = "app/static/img/logo.png"

    # Fondo de cabecera
    p.setFillColorRGB(0.18, 0.49, 0.20)  # Verde corporativo
    p.rect(0, height - 1.3 * inch, width, 1.3 * inch, fill=True, stroke=False)

    # Logo (opcional)
    try:
        p.drawImage(logo_path, 0.6 * inch, height - 1.1 * inch, width=70, height=70, mask='auto')
    except:
        pass  # si no hay logo, sigue igual

    # Texto de cabecera
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 13)
    p.drawString(1.8 * inch, height - 0.7 * inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L.")

    p.setFont("Helvetica", 9)
    p.drawString(1.8 * inch, height - 0.9 * inch, "RUC: 20610415726")
    p.drawString(1.8 * inch, height - 1.05 * inch, "Dirección: Otr. Sector Conchuc Lote 7, Caraz, Huaylas, Ancash, Perú")

    # Línea separadora
    p.setStrokeColorRGB(0.18, 0.49, 0.20)
    p.setLineWidth(2)
    p.line(0.6 * inch, height - 1.4 * inch, width - 0.6 * inch, height - 1.4 * inch)

    # === TÍTULO DEL REPORTE ===
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(1 * inch, height - 1.8 * inch, "📋 Reporte de Usuarios del Sistema")

    # Fecha de generación
    p.setFont("Helvetica", 9)
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    p.drawRightString(width - 0.6 * inch, height - 1.8 * inch, f"Generado el: {fecha_actual}")

    # === ENCABEZADO DE TABLA ===
    y = height - 2.2 * inch
    header_x = 0.6 * inch
    header_w = width - 1.2 * inch
    header_h = 0.3 * inch

    columnas = ["ID", "Nombre", "DNI", "Rol", "Email", "Estado", "Creación"]
    # Asignamos proporciones relativas (suma ≈ 1.0)
    proporciones = [0.07, 0.18, 0.12, 0.15, 0.26, 0.10, 0.12]
    anchos = [header_w * p for p in proporciones]

    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(header_x, y - 0.2 * inch, header_w, header_h, fill=True, stroke=False)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 9)

    x_actual = header_x
    for i, col in enumerate(columnas):
        p.drawCentredString(x_actual + anchos[i] / 2, y - 0.05 * inch, col)
        x_actual += anchos[i]

    # === CUERPO DE TABLA ===
    p.setFont("Helvetica", 8)
    p.setFillColor(colors.black)
    y -= 0.35 * inch

    for u in usuarios:
        if y < 1 * inch:
            p.showPage()
            y = height - 1 * inch

            # Repetir encabezado
            p.setFont("Helvetica-Bold", 9)
            p.setFillColorRGB(0.18, 0.49, 0.20)
            p.rect(header_x, y - 0.2 * inch, header_w, header_h, fill=True, stroke=False)
            p.setFillColor(colors.white)
            x_actual = header_x
            for i, col in enumerate(columnas):
                p.drawCentredString(x_actual + anchos[i] / 2, y - 0.05 * inch, col)
                x_actual += anchos[i]
            y -= 0.35 * inch
            p.setFont("Helvetica", 8)
            p.setFillColor(colors.black)

        estado_texto = "Activo" if u.get("estado") == 1 else "Inactivo"
        datos = [
            str(u.get("id_usuario", "")),
            (u.get("nombre_completo") or "")[:20],
            str(u.get("dni", "")),
            str(u.get("nombre_rol", "")),
            (u.get("email") or "")[:30],
            estado_texto,
            str(u.get("fecha_creacion") or "")
        ]

        x_actual = header_x
        
        for i, dato in enumerate(datos):
            # Centramos cada texto bajo su encabezado
            p.drawCentredString(x_actual + anchos[i] / 2, y, dato)
            x_actual += anchos[i]

        y -= 0.25 * inch

    # Pie de página
    p.setFont("Helvetica-Oblique", 8)
    p.setFillColor(colors.gray)
    p.drawCentredString(width / 2, 0.5 * inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L. - Sistema de Gestión © 2025")

    p.save()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_usuarios.pdf',
        mimetype='application/pdf'
    )