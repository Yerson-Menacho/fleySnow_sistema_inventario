from flask import Blueprint, render_template, request, jsonify, send_file
from io import BytesIO
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from babel.dates import format_date

from app.helpers import buscar_agricultores_json
from app.models import get_db_connection
from app.models.agricultores_model import listar_agricultores
from app.models.reportes_model import obtener_reporte_materia_prima
from app.routes.reportes.reportes_usuarios import reportes_bp
from app.models.recepcion_model import listar_recepciones
from app.utils.auth import permiso_de_admin

# ==============================
# 📋 Vista principal - Reporte de Recepciones
# ==============================
@reportes_bp.route('/reportes/recepciones')
@permiso_de_admin('recepciones.ver')
def vista_reporte_recepciones():
    return render_template('reportes/reporte_recepcion.html')


# ==============================
# 📊 Data vía AJAX - Reporte de Recepciones
# ==============================
@reportes_bp.route('/reportes/recepciones/data')
@permiso_de_admin('recepciones.ver')
def data_reporte_recepciones():
    p_q = request.args.get("q", None)
    p_estado = request.args.get("estado", None)
    p_fecha_inicio = request.args.get("fecha_inicio", None)
    p_fecha_fin = request.args.get("fecha_fin", None)
    p_id_agricultor = request.args.get("id_agricultor", None) 

    data_raw = listar_recepciones(p_q, p_estado, p_fecha_inicio, p_fecha_fin, p_id_agricultor) 

    data = []
    for r in data_raw:
        data.append({
            "id_recepcion": r.get("id_recepcion"),
            "codigo_recepcion": r.get("codigo_recepcion"),
            "agricultor": r.get("agricultor"),
            "fecha_recepcion": r.get("fecha_recepcion"),
            "peso_bruto_total": float(r.get("peso_bruto_total", 0)),
            "jabas_entregadas": int(r.get("jabas_entregadas", 0)),
            "peso_neto_total": float(r.get("peso_neto_total", 0)),
            "jabas_enviadas": int(r.get("jabas_enviadas", 0)),
            "saldo_jabas": int(r.get("saldo_jabas", 0)),
            "estado": r.get("estado"),
        })

    return jsonify(data)

@reportes_bp.route('/buscar_agricultores')
def buscar_agricultores():
    return buscar_agricultores_json()

# ==============================
# 📈 Exportar Excel - Reporte de Recepciones
# ==============================
@reportes_bp.route('/reportes/recepciones/excel')
@permiso_de_admin('recepciones.ver')
def exportar_reporte_recepciones_excel():
    recepciones = listar_recepciones()

    df = pd.DataFrame(recepciones)

    if not df.empty:
        df = df.rename(columns={
            "id_recepcion": "ID",
            "codigo_recepcion": "Código",
            "agricultor": "Agricultor",
            "fecha_recepcion": "Fecha",
            "peso_bruto_total": "Peso Bruto Total (kg)",
            "jabas_entregadas": "Jabas Entregadas",
            "peso_neto_total": "Peso Neto Total (kg)",
            "jabas_enviadas": "Jabas Enviadas",
            "saldo_jabas": "Saldo de Jabas",
            "estado": "Estado"
        })

        df = df[[
            "ID", "Código", "Agricultor", "Fecha",
            "Peso Bruto Total (kg)", "Jabas Enviadas", "Jabas Entregadas",
            "Peso Neto Total (kg)", "Saldo de Jabas", "Estado"
        ]]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Recepciones')
        wb = writer.book
        ws = wb.active

        # === CABECERA EMPRESARIAL ===
        ws.insert_rows(1, 6)
        ws.merge_cells('A1:J1')
        ws.merge_cells('A2:J2')
        ws.merge_cells('A3:J3')
        ws.merge_cells('A4:J4')
        ws.merge_cells('A5:J5')

        ws['A1'] = "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L."
        ws['A2'] = "RUC: 20610415726"
        ws['A3'] = "Dirección: Otr. Sector Conchuc Lote 7, Caraz, Huaylas, Ancash, Perú"
        ws['A4'] = "📑 Reporte de Recepciones de Materia Prima"
        ws['A5'] = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        for i in range(1, 6):
            cell = ws[f"A{i}"]
            cell.font = Font(bold=True, size=12 if i == 1 else 10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # === ENCABEZADO ===
        header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[7]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # === CUERPO ===
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # === AUTO AJUSTE ===
        for i, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            column = get_column_letter(i)
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_recepciones.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ==============================
# 📄 Exportar PDF - Reporte de Recepciones
# ==============================
@reportes_bp.route('/reportes/recepciones/pdf')
@permiso_de_admin('recepciones.ver')
def exportar_reporte_recepciones_pdf():
    recepciones = listar_recepciones()
    output = BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    width, height = letter

    # === CABECERA EMPRESARIAL ===
    logo_path = "app/static/img/logo.png"

    # Fondo superior verde
    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(0, height - 1.3 * inch, width, 1.3 * inch, fill=True, stroke=False)

    # Logo
    try:
        p.drawImage(logo_path, 0.6 * inch, height - 1.1 * inch, width=70, height=70, mask='auto')
    except:
        pass

    # Texto empresa
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 13)
    p.drawString(1.8 * inch, height - 0.7 * inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L.")
    p.setFont("Helvetica", 9)
    p.drawString(1.8 * inch, height - 0.9 * inch, "RUC: 20610415726")
    p.drawString(1.8 * inch, height - 1.05 * inch, "Dirección: Otr. Sector Conchuc Lote 7, Caraz, Huaylas, Ancash, Perú")

    # Línea separadora
    p.setStrokeColorRGB(0.18, 0.49, 0.20)
    p.setLineWidth(2)
    p.line(0.6 * inch, height - 1.4 * inch, width - 0.6 * inch, height - 1.4 * inch)

    # Título
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(1 * inch, height - 1.8 * inch, "📑 Reporte de Recepciones de Materia Prima")

    # Fecha
    p.setFont("Helvetica", 9)
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    p.drawRightString(width - 0.6 * inch, height - 1.8 * inch, f"Generado el: {fecha_actual}")

    # === ENCABEZADO DE TABLA ===
    y = height - 2.2 * inch
    header_x = 0.4 * inch
    header_w = width - 0.8 * inch
    header_h = 0.3 * inch

    columnas = ["Código", "Agricultor", "Fecha", "Bruto (kg)", "Net (kg)", "J.Env", "J.Ent", "Saldo", "Estado"]
    proporciones = [0.12, 0.2, 0.1, 0.1, 0.1, 0.08, 0.08, 0.08, 0.08]
    anchos = [header_w * p for p in proporciones]

    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(header_x, y - 0.2 * inch, header_w, header_h, fill=True, stroke=False)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 8)

    x_actual = header_x
    for i, col in enumerate(columnas):
        p.drawCentredString(x_actual + anchos[i] / 2, y - 0.05 * inch, col)
        x_actual += anchos[i]

    # === CUERPO ===
    p.setFont("Helvetica", 7.5)
    p.setFillColor(colors.black)
    y -= 0.35 * inch

    for r in recepciones:
        if y < 1 * inch:
            p.showPage()
            y = height - 1 * inch
            p.setFont("Helvetica", 7.5)

        datos = [
            str(r.get("codigo_recepcion", "")),
            str(r.get("agricultor", ""))[:18],
            str(r.get("fecha_recepcion", ""))[:10],
            str(r.get("peso_bruto_total", "")),
            str(r.get("peso_neto_total", "")),
            str(r.get("jabas_enviadas", "")),
            str(r.get("jabas_entregadas", "")),
            str(r.get("saldo_jabas", "")),
            str(r.get("estado", "")),
        ]

        x_actual = header_x
        for i, dato in enumerate(datos):
            p.drawCentredString(x_actual + anchos[i] / 2, y, dato)
            x_actual += anchos[i]
        y -= 0.25 * inch

    p.setFont("Helvetica-Oblique", 8)
    p.setFillColor(colors.gray)
    p.drawCentredString(width / 2, 0.5 * inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L. - Sistema de Gestión © 2025")

    p.save()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_recepciones.pdf',
        mimetype='application/pdf'
    )


# --- Vista principal ---
@reportes_bp.route('/reportes/materia_prima')
@permiso_de_admin('recepciones.ver')
def vista_reporte_materia_prima():
    return render_template('reportes/reporte_total_materia_prima.html')


# --- Datos del reporte ---
@reportes_bp.route('/reportes/materia_prima/data')
@permiso_de_admin('recepciones.ver')
def data_reporte_materia_prima():
    p_fecha_inicio = request.args.get("fecha_inicio")
    p_fecha_fin = request.args.get("fecha_fin")
    p_id_variedad = request.args.get("id_variedad")
    p_id_agricultor = request.args.get("id_agricultor")
    p_agrupacion = request.args.get("agrupacion", "dia")

    data_raw = obtener_reporte_materia_prima(
        fecha_desde=p_fecha_inicio,
        fecha_hasta=p_fecha_fin,
        id_variedad=p_id_variedad,
        id_agricultor=p_id_agricultor,
        agrupacion=p_agrupacion
    )

    data = []
    for r in data_raw:
        periodo = r.get("periodo")

        if p_agrupacion == "dia":
            try:
                # Detectar si viene como Wed, 22 Oct 2025 00:00:00 GMT
                if isinstance(periodo, str) and "GMT" in periodo:
                    periodo = datetime.strptime(periodo.replace("GMT", "").strip(), "%a, %d %b %Y %H:%M:%S")
                elif isinstance(periodo, str):
                    periodo = datetime.strptime(periodo[:10], "%Y-%m-%d")
                periodo = periodo.strftime("%d/%m/%Y")
            except Exception:
                periodo = str(periodo)

        elif p_agrupacion == "semana":
            try:
                # Aseguramos formato 2025-W43
                periodo = periodo.replace(" ", " ")
                if "-W" in periodo:
                    anio, semana = periodo.split("-W")
                    periodo = f"Semana {int(semana)} del {anio}"
            except Exception:
                periodo = str(periodo)

        elif p_agrupacion == "mes":
            try:
                periodo = datetime.strptime(periodo, "%Y-%m").strftime("%Y-%m")
            except Exception:
                periodo = str(periodo)

        data.append({
            "periodo": periodo,
            "variedad": r.get("variedad"),
            "total_bruto": float(r.get("total_bruto", 0)),
            "total_jabas": int(r.get("total_jabas", 0)),
            "total_neto": float(r.get("total_bruto", 0)) - (int(r.get("total_jabas", 0)) * 2),
            "recepciones_count": int(r.get("recepciones_count", 0)),
        })

    return jsonify(data)

@reportes_bp.route('/reportes/materia_prima/pdf')
@permiso_de_admin('recepciones.ver')
def exportar_reporte_materia_prima_pdf():
    # Filtros
    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")
    id_variedad = request.args.get("id_variedad")
    id_agricultor = request.args.get("id_agricultor")
    agrupacion = request.args.get("agrupacion", "dia")

    # Datos desde el SP
    data = obtener_reporte_materia_prima(
        fecha_desde=fecha_inicio,
        fecha_hasta=fecha_fin,
        id_variedad=id_variedad,
        id_agricultor=id_agricultor,
        agrupacion=agrupacion
    )

    output = BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    width, height = letter

    # === CABECERA EMPRESARIAL ===
    logo_path = "app/static/img/logo.png"
    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(0, height - 1.3 * inch, width, 1.3 * inch, fill=True, stroke=False)

    # Logo
    try:
        p.drawImage(logo_path, 0.6 * inch, height - 1.1 * inch, width=70, height=70, mask='auto')
    except:
        pass

    # Texto empresa
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 13)
    p.drawString(1.8 * inch, height - 0.7 * inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L.")
    p.setFont("Helvetica", 9)
    p.drawString(1.8 * inch, height - 0.9 * inch, "RUC: 20610415726")
    p.drawString(1.8 * inch, height - 1.05 * inch, "Dirección: Otr. Sector Conchuc Lote 7, Caraz, Huaylas, Ancash, Perú")

    # Línea separadora
    p.setStrokeColorRGB(0.18, 0.49, 0.20)
    p.setLineWidth(2)
    p.line(0.6 * inch, height - 1.4 * inch, width - 0.6 * inch, height - 1.4 * inch)

    # === Título y Fecha ===
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(1 * inch, height - 1.8 * inch, "📑 Reporte de Materia Prima")

    p.setFont("Helvetica", 9)
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    p.drawRightString(width - 0.6 * inch, height - 1.8 * inch, f"Generado el: {fecha_actual}")

    # === ENCABEZADO DE TABLA ===
    y = height - 2.2 * inch
    header_x = 0.4 * inch
    header_w = width - 0.8 * inch
    header_h = 0.3 * inch

    columnas = ["Periodo", "Variedad", "Peso Bruto (kg)", "Cantidad Jabas", "Peso Neto (kg)", "N° Recepciones"]
    proporciones = [0.15, 0.2, 0.18, 0.15, 0.15, 0.12]
    anchos = [header_w * p for p in proporciones]

    p.setFillColorRGB(0.18, 0.49, 0.20)
    p.rect(header_x, y - 0.2 * inch, header_w, header_h, fill=True, stroke=False)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 8)

    x_actual = header_x
    for i, col in enumerate(columnas):
        p.drawCentredString(x_actual + anchos[i]/2, y - 0.05*inch, col)
        x_actual += anchos[i]

    # === CUERPO DE TABLA ===
    p.setFont("Helvetica", 7.5)
    p.setFillColor(colors.black)
    y -= 0.35 * inch

    for r in data:
        if y < 1 * inch:
            p.showPage()
            y = height - 1 * inch
            p.setFont("Helvetica", 7.5)

        total_bruto = float(r.get("total_bruto", 0))
        total_jabas = int(r.get("total_jabas", 0))
        total_neto = total_bruto - (total_jabas * 2)

        datos = [
            str(r.get("periodo", "")),
            str(r.get("variedad", ""))[:18],
            f"{total_bruto:,.2f}",
            str(total_jabas),
            f"{total_neto:,.2f}",
            str(r.get("recepciones_count", 0)),
        ]

        x_actual = header_x
        for i, dato in enumerate(datos):
            p.drawCentredString(x_actual + anchos[i]/2, y, dato)
            x_actual += anchos[i]
        y -= 0.25 * inch

    # === PIE DE PÁGINA ===
    p.setFont("Helvetica-Oblique", 8)
    p.setFillColor(colors.gray)
    p.drawCentredString(width / 2, 0.5 * inch, "FLEY & SNOW INTERNATIONAL PRODUCE E.I.R.L. - Sistema de Gestión © 2025")

    p.save()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_materia_prima.pdf',
        mimetype='application/pdf'
    )
